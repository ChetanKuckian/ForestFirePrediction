import pandas
from openpyxl import load_workbook
import xlwt
import xlsxwriter

headings = ['Drought Factor','Temperature', 'Relative Humidity', 'Wind', 'FFDI']

def createSheet(bookname):
    workbook=xlwt.Workbook(bookname)
    workbook.add_sheet('Low')
    workbook.add_sheet('Medium')
    workbook.add_sheet('High')
    workbook.save(bookname)
    tp(bookname)


def tp(bookname):
    dataFrame = pandas.DataFrame([],columns = headings)
    #book = load_workbook(bookname)
    writer = pandas.ExcelWriter(bookname, engine='xlsxwriter')
    #writer.book = book
    dataFrame.to_excel(writer,sheet_name='Low',index=False)
    dataFrame.to_excel(writer,sheet_name='Medium',index=False)
    dataFrame.to_excel(writer,sheet_name='High',index=False)
    writer.save()

def appendToExcel(bookname,givenList):
    dataFrame = pandas.DataFrame(givenList,columns = headings)
    FfdiLow = dataFrame[dataFrame.FFDI <= 10]
    FfdiMedium = dataFrame[dataFrame.FFDI > 10]
    FfdiMedium = FfdiMedium[FfdiMedium.FFDI <=20]
    FfdiHigh = dataFrame[dataFrame.FFDI > 20]

    book = load_workbook(bookname)
    writer = pandas.ExcelWriter(bookname, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        if sheetname == 'Low':
            FfdiLow.to_excel(writer,sheet_name = sheetname,startrow= writer.sheets[sheetname].max_row ,index=False,header=False)
        elif sheetname == 'Medium':    
            FfdiMedium.to_excel(writer,sheet_name = sheetname,startrow= writer.sheets[sheetname].max_row ,index=False,header=False)
        else:
            FfdiHigh.to_excel(writer,sheet_name = sheetname,startrow= writer.sheets[sheetname].max_row ,index=False,header=False)
    writer.save()
@supress 
def createBestIndividualExcel(bookname,heading):
    workbook=xlwt.Workbook(bookname)
    workbook.add_sheet('Sheet1')
    workbook.save(bookname)
    dataFrame = pandas.DataFrame([],columns = heading)
    writer = pandas.ExcelWriter(bookname, engine='xlsxwriter')
    dataFrame.to_excel(writer,sheet_name='Sheet1',index=False)
    writer.save()



def bestIndividualToExcel(bookname,individual,heading):
    dataFrame = pandas.DataFrame([individual],columns = heading)
    book = load_workbook(bookname)
    writer = pandas.ExcelWriter(bookname, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        if sheetname == 'Sheet1':
            dataFrame.to_excel(writer,sheet_name = sheetname,startrow= writer.sheets[sheetname].max_row ,index=False,header=False)
    writer.save()
